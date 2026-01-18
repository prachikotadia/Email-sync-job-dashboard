"""
Export Service - Production-grade export functionality
Generates CSV, Excel, JSON, and PDF exports from application data
"""
import csv
import io
import json
import logging
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from app.database import Application, User
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)

# Category mapping
CATEGORY_MAP = {
    "ALL": None,
    "APPLIED": "applied",
    "REJECTED": "rejected",
    "INTERVIEW": "interview",
    "OFFER": "offer",
    "GHOSTED": "ghosted",
}


def get_applications_for_export(
    db: Session,
    user_id: int,
    category: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
) -> List[Application]:
    """
    Query applications with filters
    Returns list of Application objects matching criteria
    """
    query = db.query(Application).filter(Application.user_id == user_id)

    # Category filter
    if category and category != "ALL":
        category_lower = CATEGORY_MAP.get(category.upper(), category.lower())
        if category_lower:
            query = query.filter(Application.category == category_lower)

    # Date range filter
    if date_from:
        query = query.filter(Application.received_at >= date_from)
    if date_to:
        # Include entire day
        date_to_end = date_to.replace(hour=23, minute=59, second=59)
        query = query.filter(Application.received_at <= date_to_end)

    # Order by received_at descending
    query = query.order_by(Application.received_at.desc())

    applications = query.all()
    logger.info(f"Export query: user_id={user_id}, category={category}, count={len(applications)}")
    return applications


def export_to_csv(
    applications: List[Application],
    fields: List[str],
    user_email: str,
) -> bytes:
    """
    Generate CSV export
    Returns UTF-8 encoded CSV bytes
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write headers
    headers = []
    field_map = {
        "company_name": "Company Name",
        "category": "Category",
        "received_at": "Received Date",
        "last_updated": "Last Updated",
        "source_email": "Source Email",
        "gmail_message_id": "Gmail Message ID",
    }
    for field in fields:
        if field in field_map:
            headers.append(field_map[field])
    writer.writerow(headers)

    # Write data rows
    for app in applications:
        row = []
        for field in fields:
            if field == "company_name":
                row.append(app.company_name or "")
            elif field == "category":
                row.append(app.category or "")
            elif field == "received_at":
                row.append(app.received_at.isoformat() if app.received_at else "")
            elif field == "last_updated":
                row.append(app.last_updated.isoformat() if app.last_updated else "")
            elif field == "source_email":
                row.append(app.from_email or "")
            elif field == "gmail_message_id":
                row.append(app.gmail_message_id or "")
        writer.writerow(row)

    csv_string = output.getvalue()
    output.close()
    return csv_string.encode("utf-8-sig")  # UTF-8 with BOM for Excel compatibility


def export_to_excel(
    applications: List[Application],
    fields: List[str],
    user_email: str,
) -> bytes:
    """
    Generate Excel (.xlsx) export
    Returns Excel file bytes
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Field mapping
    field_map = {
        "company_name": "Company Name",
        "category": "Category",
        "received_at": "Received Date",
        "last_updated": "Last Updated",
        "source_email": "Source Email",
        "gmail_message_id": "Gmail Message ID",
    }

    # Write headers with styling
    headers = [field_map.get(field, field) for field in fields]
    header_row = ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for app in applications:
        row = []
        for field in fields:
            if field == "company_name":
                row.append(app.company_name or "")
            elif field == "category":
                row.append(app.category or "")
            elif field == "received_at":
                row.append(app.received_at.isoformat() if app.received_at else "")
            elif field == "last_updated":
                row.append(app.last_updated.isoformat() if app.last_updated else "")
            elif field == "source_email":
                row.append(app.from_email or "")
            elif field == "gmail_message_id":
                row.append(app.gmail_message_id or "")
        ws.append(row)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_to_json(
    applications: List[Application],
    fields: List[str],
    user_email: str,
) -> bytes:
    """
    Generate JSON export
    Returns JSON bytes (array of objects, no wrapper)
    """
    data = []
    for app in applications:
        obj = {}
        for field in fields:
            if field == "company_name":
                obj["company_name"] = app.company_name
            elif field == "category":
                obj["category"] = app.category
            elif field == "received_at":
                obj["received_at"] = app.received_at.isoformat() if app.received_at else None
            elif field == "last_updated":
                obj["last_updated"] = app.last_updated.isoformat() if app.last_updated else None
            elif field == "source_email":
                obj["source_email"] = app.from_email
            elif field == "gmail_message_id":
                obj["gmail_message_id"] = app.gmail_message_id
        data.append(obj)

    json_string = json.dumps(data, indent=2, default=str)
    return json_string.encode("utf-8")


def export_to_pdf(
    applications: List[Application],
    fields: List[str],
    user_email: str,
    category: str,
    date_from: Optional[datetime],
    date_to: Optional[datetime],
) -> bytes:
    """
    Generate PDF summary report
    Returns PDF bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a2234"),
        spaceAfter=12,
    )
    story.append(Paragraph("Job Application Export Report", title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Metadata
    meta_style = styles["Normal"]
    story.append(Paragraph(f"<b>User:</b> {user_email}", meta_style))
    story.append(Paragraph(
        f"<b>Export Date:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        meta_style
    ))
    story.append(Spacer(1, 0.2 * inch))

    # Filters applied
    filters = []
    if category and category != "ALL":
        filters.append(f"Category: {category}")
    if date_from:
        filters.append(f"From: {date_from.strftime('%Y-%m-%d')}")
    if date_to:
        filters.append(f"To: {date_to.strftime('%Y-%m-%d')}")
    if filters:
        story.append(Paragraph(f"<b>Filters:</b> {', '.join(filters)}", meta_style))
    story.append(Spacer(1, 0.2 * inch))

    # Summary statistics
    story.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
    total = len(applications)
    category_counts = {}
    for app in applications:
        cat = app.category or "Unknown"
        category_counts[cat] = category_counts.get(cat, 0) + 1

    summary_data = [["Metric", "Count"]]
    summary_data.append(["Total Applications", str(total)])
    for cat, count in sorted(category_counts.items()):
        summary_data.append([cat.capitalize(), str(count)])

    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * inch))

    # Applications table (if not too many)
    if total > 0 and total <= 100:  # Only show table for reasonable number of records
        story.append(Paragraph("<b>Applications</b>", styles["Heading2"]))
        
        # Prepare table data
        field_map = {
            "company_name": "Company",
            "category": "Category",
            "received_at": "Received",
            "last_updated": "Updated",
            "source_email": "From",
            "gmail_message_id": "Message ID",
        }
        
        table_headers = [field_map.get(field, field) for field in fields]
        table_data = [table_headers]
        
        for app in applications[:50]:  # Limit to 50 rows in PDF
            row = []
            for field in fields:
                if field == "company_name":
                    row.append(app.company_name or "")
                elif field == "category":
                    row.append(app.category or "")
                elif field == "received_at":
                    row.append(app.received_at.strftime("%Y-%m-%d") if app.received_at else "")
                elif field == "last_updated":
                    row.append(app.last_updated.strftime("%Y-%m-%d") if app.last_updated else "")
                elif field == "source_email":
                    row.append(app.from_email or "")
                elif field == "gmail_message_id":
                    row.append(app.gmail_message_id[:20] + "..." if app.gmail_message_id and len(app.gmail_message_id) > 20 else (app.gmail_message_id or ""))
            table_data.append(row)
        
        if total > 50:
            table_data.append(["", f"... and {total - 50} more records"])
        
        # Create table
        table = Table(table_data, colWidths=[1.5 * inch] * len(fields))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.beige]),
                ]
            )
        )
        story.append(table)
    elif total > 100:
        story.append(Paragraph(
            f"<i>Table view omitted (too many records: {total}). Use CSV/Excel for full data.</i>",
            meta_style
        ))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def generate_export(
    db: Session,
    user_id: int,
    user_email: str,
    format_type: str,
    category: str,
    date_from: Optional[datetime],
    date_to: Optional[datetime],
    fields: List[str],
) -> Tuple[bytes, str, str]:
    """
    Main export function
    Returns: (file_bytes, mime_type, filename)
    """
    # Validate format
    valid_formats = ["csv", "xlsx", "json", "pdf"]
    if format_type.lower() not in valid_formats:
        raise ValueError(f"Invalid format: {format_type}. Must be one of {valid_formats}")

    # Validate fields
    valid_fields = [
        "company_name",
        "category",
        "received_at",
        "last_updated",
        "source_email",
        "gmail_message_id",
    ]
    if not fields:
        raise ValueError("At least one field must be selected")
    for field in fields:
        if field not in valid_fields:
            raise ValueError(f"Invalid field: {field}")

    # Query applications
    applications = get_applications_for_export(db, user_id, category, date_from, date_to)
    count = len(applications)

    logger.info(
        f"Export: user={user_email}, format={format_type}, category={category}, "
        f"date_from={date_from}, date_to={date_to}, records={count}"
    )

    # Generate file based on format
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    
    if format_type.lower() == "csv":
        file_bytes = export_to_csv(applications, fields, user_email)
        mime_type = "text/csv"
        filename = f"applications_export_{timestamp}.csv"
    elif format_type.lower() == "xlsx":
        file_bytes = export_to_excel(applications, fields, user_email)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"applications_export_{timestamp}.xlsx"
    elif format_type.lower() == "json":
        file_bytes = export_to_json(applications, fields, user_email)
        mime_type = "application/json"
        filename = f"applications_export_{timestamp}.json"
    elif format_type.lower() == "pdf":
        file_bytes = export_to_pdf(applications, fields, user_email, category, date_from, date_to)
        mime_type = "application/pdf"
        filename = f"applications_export_{timestamp}.pdf"
    else:
        raise ValueError(f"Unsupported format: {format_type}")

    logger.info(f"Export SUCCESS: user={user_email}, format={format_type}, records={count}, filename={filename}")
    return file_bytes, mime_type, filename
